"""항목별 엑셀 입출력.

- GET  /excel/entities            지원 항목 목록
- GET  /excel/template/{entity}   입력 양식(xlsx) 다운로드 — 헤더+예시행+작성안내 시트
- GET  /excel/export/{entity}     현재 DB 데이터 내보내기(xlsx)
- POST /excel/import/{entity}     양식에 채운 데이터 업로드 → DB 반영 (생성/갱신/오류 리포트)

코드 참조(model_code/site_code/asset_code/part_no)는 임포트 시 자동으로 ID 해석.
upsert 키가 있는 항목은 기존 행을 갱신, 없으면 신규 생성. 이력성 항목(BM/이슈/L&L)은 추가만.
"""
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db

router = APIRouter(prefix="/excel", tags=["excel"])


# ───────────────────────── 변환 유틸 ─────────────────────────

def _s(v):  # str
    return str(v).strip() if v is not None else ""


def _f(v):  # float | None
    if v is None or str(v).strip() == "":
        return None
    return float(str(v).replace(",", ""))


def _i(v):  # int | None
    f = _f(v)
    return int(f) if f is not None else None


def _b(v):  # bool
    return _s(v).upper() in ("Y", "YES", "TRUE", "1", "O")


def _d(v):  # date | None
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return date.fromisoformat(str(v).strip()[:10])


def _dt(v):  # datetime | None
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip().replace(" ", "T")
    return datetime.fromisoformat(s[:19])


class Ref:
    """코드 → ID 해석기."""
    def __init__(self, db: Session):
        self.db = db

    def model(self, code, required=True):
        if not _s(code):
            if required:
                raise ValueError("model_code 누락")
            return None
        m = self.db.query(models.EquipmentModel).filter(
            models.EquipmentModel.code == _s(code)).first()
        if not m:
            raise ValueError(f"모델 코드 없음: {code}")
        return m.id

    def site(self, code, required=True):
        if not _s(code):
            if required:
                raise ValueError("site_code 누락")
            return None
        s = self.db.query(models.Site).filter(models.Site.code == _s(code)).first()
        if not s:
            raise ValueError(f"사이트 코드 없음: {code}")
        return s.id

    def equipment(self, code, required=True):
        if not _s(code):
            if required:
                raise ValueError("asset_code 누락")
            return None
        e = self.db.query(models.Equipment).filter(
            models.Equipment.asset_code == _s(code)).first()
        if not e:
            raise ValueError(f"설비 코드 없음: {code}")
        return e.id

    def part(self, no, required=True):
        if not _s(no):
            if required:
                raise ValueError("part_no 누락")
            return None
        p = self.db.query(models.Part).filter(models.Part.part_no == _s(no)).first()
        if not p:
            raise ValueError(f"파츠 번호 없음: {no}")
        return p.id


# ───────────────────────── 항목 정의 ─────────────────────────
# columns: (헤더, 필수여부, 설명, 예시)
# importer(db, ref, row:dict) -> ("created"|"updated", None) — 행 단위 반영
# exporter(db) -> list[list] — 현재 데이터 행들

SPECS: dict[str, dict] = {}


def spec(key, label, desc, columns, importer, exporter, insert_only=False):
    SPECS[key] = {"label": label, "desc": desc, "columns": columns,
                  "importer": importer, "exporter": exporter, "insert_only": insert_only}


# 1) 설비
def _imp_equipment(db, ref, r):
    eq = db.query(models.Equipment).filter(
        models.Equipment.asset_code == _s(r["asset_code"])).first()
    data = dict(model_id=ref.model(r["model_code"]), site_id=ref.site(r["site_code"]),
                line=_s(r.get("line")), status=_s(r.get("status")) or "DR",
                install_date=_d(r.get("install_date")),
                annual_run_hours=_f(r.get("annual_run_hours")) or 6000)
    if eq:
        for k, v in data.items():
            setattr(eq, k, v)
        return "updated"
    db.add(models.Equipment(asset_code=_s(r["asset_code"]), **data))
    return "created"


spec("equipments", "설비 마스터", "asset_code 기준 생성/갱신",
     [("asset_code", True, "설비 자산코드 (고유)", "STK-2000-004"),
      ("model_code", True, "설비 모델 코드", "STK-2000"),
      ("site_code", True, "사이트 코드", "KR1"),
      ("line", False, "라인", "A4"),
      ("status", False, "DR/FAB/SETUP/RUN/PM/BM/STOP/SCRAP", "RUN"),
      ("install_date", False, "설치일 YYYY-MM-DD", "2025-01-15"),
      ("annual_run_hours", False, "연간 가동시간", "7000")],
     _imp_equipment,
     lambda db: [[e.asset_code, e.model.code, e.site.code, e.line, e.status,
                  e.install_date, e.annual_run_hours]
                 for e in db.query(models.Equipment).all()])


# 2) 파츠
def _imp_part(db, ref, r):
    p = db.query(models.Part).filter(models.Part.part_no == _s(r["part_no"])).first()
    data = dict(name=_s(r["name"]), category=_s(r.get("category")), maker=_s(r.get("maker")),
                unit_price=_f(r.get("unit_price")) or 0,
                lead_time_days=_i(r.get("lead_time_days")) or 30,
                mtbf_hours=_f(r.get("mtbf_hours")),
                current_stock=_i(r.get("current_stock")) or 0,
                min_stock=_i(r.get("min_stock")) or 0)
    if p:
        for k, v in data.items():
            setattr(p, k, v)
        return "updated"
    db.add(models.Part(part_no=_s(r["part_no"]), **data))
    return "created"


spec("parts", "스페어 파츠", "part_no 기준 생성/갱신 (재고 포함)",
     [("part_no", True, "파츠 번호 (고유)", "BRG-6312ZZ"),
      ("name", True, "품명", "베어링 6312ZZ"),
      ("category", False, "분류", "구동"),
      ("maker", False, "메이커", "NSK"),
      ("unit_price", False, "단가(원)", "52000"),
      ("lead_time_days", False, "리드타임(일)", "14"),
      ("mtbf_hours", False, "MTBF(시간)", "40000"),
      ("current_stock", False, "현재고", "4"),
      ("min_stock", False, "최소재고", "2")],
     _imp_part,
     lambda db: [[p.part_no, p.name, p.category, p.maker, p.unit_price,
                  p.lead_time_days, p.mtbf_hours, p.current_stock, p.min_stock]
                 for p in db.query(models.Part).all()])


# 3) BOM
def _imp_bom(db, ref, r):
    mid, pid = ref.model(r["model_code"]), ref.part(r["part_no"])
    b = db.query(models.ModelPartBom).filter(
        models.ModelPartBom.model_id == mid, models.ModelPartBom.part_id == pid).first()
    data = dict(qty_per_unit=_i(r.get("qty_per_unit")) or 1,
                replace_cycle_months=_i(r.get("replace_cycle_months")),
                critical=_b(r.get("critical")))
    if b:
        for k, v in data.items():
            setattr(b, k, v)
        return "updated"
    db.add(models.ModelPartBom(model_id=mid, part_id=pid, **data))
    return "created"


spec("bom", "모델별 BOM", "모델×파츠 기준 생성/갱신 — 권장재고 산출 기초",
     [("model_code", True, "설비 모델 코드", "STK-2000"),
      ("part_no", True, "파츠 번호", "BRG-6312ZZ"),
      ("qty_per_unit", False, "대당 수량", "4"),
      ("replace_cycle_months", False, "교체주기(월)", "36"),
      ("critical", False, "Critical 여부 Y/N", "Y")],
     _imp_bom,
     lambda db: [[b.model.code, b.part.part_no, b.qty_per_unit,
                  b.replace_cycle_months, "Y" if b.critical else "N"]
                 for b in db.query(models.ModelPartBom).all()])


# 4) PM 표준
def _imp_pmstd(db, ref, r):
    mid = ref.model(r["model_code"])
    it = db.query(models.PMStandardItem).filter(
        models.PMStandardItem.model_id == mid,
        models.PMStandardItem.item_no == _s(r["item_no"])).first()
    data = dict(name=_s(r["name"]), part_area=_s(r.get("part_area")),
                method=_s(r.get("method")).upper() or "VISUAL",
                criteria=_s(r.get("criteria")),
                lower_limit=_f(r.get("lower_limit")), upper_limit=_f(r.get("upper_limit")),
                unit=_s(r.get("unit")), period_days=_i(r.get("period_days")) or 90,
                vision_capable=_b(r.get("vision_capable")))
    if it:
        for k, v in data.items():
            setattr(it, k, v)
        return "updated"
    db.add(models.PMStandardItem(model_id=mid, item_no=_s(r["item_no"]), **data))
    return "created"


spec("pm_standards", "PM 표준 점검항목", "모델×항목번호 기준 생성/갱신 — 전사 공통 표준",
     [("model_code", True, "설비 모델 코드", "STK-2000"),
      ("item_no", True, "항목 번호", "PM-06"),
      ("name", True, "점검 항목명", "주행 모터 온도 점검"),
      ("part_area", False, "점검 부위", "주행부"),
      ("method", False, "VISUAL/MEASURE/VISION/REPLACE/CLEAN", "MEASURE"),
      ("criteria", False, "판정 기준 설명", "70°C 이하"),
      ("lower_limit", False, "하한", ""),
      ("upper_limit", False, "상한", "70"),
      ("unit", False, "단위", "°C"),
      ("period_days", False, "주기(일)", "90"),
      ("vision_capable", False, "비전 측정 가능 Y/N", "N")],
     _imp_pmstd,
     lambda db: [[s.model_id and db.get(models.EquipmentModel, s.model_id).code, s.item_no,
                  s.name, s.part_area, s.method, s.criteria, s.lower_limit, s.upper_limit,
                  s.unit, s.period_days, "Y" if s.vision_capable else "N"]
                 for s in db.query(models.PMStandardItem).all()])


# 5) BM 이력 (추가 전용)
def _imp_bm(db, ref, r):
    db.add(models.BMReport(
        equipment_id=ref.equipment(r["asset_code"]),
        occurred_at=_dt(r.get("occurred_at")) or datetime.utcnow(),
        symptom=_s(r["symptom"]), cause=_s(r.get("cause")), action=_s(r.get("action")),
        downtime_min=_f(r.get("downtime_min")) or 0,
        failure_part_id=ref.part(r.get("part_no"), required=False),
        status=_s(r.get("status")).upper() or "CLOSED",
        reported_by=_s(r.get("reported_by"))))
    return "created"


spec("bm_reports", "BM 고장 이력", "추가 전용 — 과거 고장 이력 일괄 적재",
     [("asset_code", True, "설비 자산코드", "STK-2000-001"),
      ("occurred_at", True, "발생일시 YYYY-MM-DD HH:MM", "2025-11-02 14:30"),
      ("symptom", True, "증상", "권상 이상 소음"),
      ("cause", False, "원인", "베어링 마모"),
      ("action", False, "조치", "베어링 교체"),
      ("downtime_min", False, "다운타임(분)", "120"),
      ("part_no", False, "고장 파츠 번호", "BRG-6310ZZ"),
      ("status", False, "OPEN/ANALYZING/FIXED/CLOSED", "CLOSED"),
      ("reported_by", False, "보고자", "최정비")],
     _imp_bm,
     lambda db: [[b.equipment.asset_code, b.occurred_at, b.symptom, b.cause, b.action,
                  b.downtime_min,
                  b.failure_part_id and db.get(models.Part, b.failure_part_id).part_no,
                  b.status, b.reported_by]
                 for b in db.query(models.BMReport).all()],
     insert_only=True)


# 6) 이슈 (추가 전용)
def _imp_issue(db, ref, r):
    sev = _s(r.get("severity")).upper() or "MID"
    dom = _s(r["domain"]).upper()
    db.add(models.Issue(
        equipment_id=ref.equipment(r.get("asset_code"), required=False),
        phase=_s(r.get("phase")).upper() or "PRODUCTION", domain=dom,
        severity="HIGH" if dom == "SAFETY" else sev,
        title=_s(r["title"]), description=_s(r.get("description")),
        owner=_s(r.get("owner")), status=_s(r.get("status")).upper() or "OPEN"))
    return "created"


spec("issues", "이슈", "추가 전용 — 안전 도메인은 자동 HIGH",
     [("title", True, "제목", "합류부 인터락 검토 필요"),
      ("domain", True, "MECH/ELEC/CONTROL/INTERLOCK/CIM/MCS/RTD/SAFETY/ETC", "INTERLOCK"),
      ("severity", False, "HIGH/MID/LOW", "HIGH"),
      ("phase", False, "INVEST/FABRICATION/SETUP/PRODUCTION", "SETUP"),
      ("asset_code", False, "설비 자산코드", "CNV-B12-001"),
      ("description", False, "상세", ""),
      ("owner", False, "담당", "이설치"),
      ("status", False, "OPEN/IN_PROGRESS/CLOSED", "OPEN")],
     _imp_issue,
     lambda db: [[i.title, i.domain, i.severity, i.phase,
                  i.equipment.asset_code if i.equipment else "",
                  i.description, i.owner, i.status]
                 for i in db.query(models.Issue).all()],
     insert_only=True)


# 7) FDC 센서
def _imp_sensor(db, ref, r):
    eid = ref.equipment(r["asset_code"])
    s = db.query(models.FDCSensor).filter(
        models.FDCSensor.equipment_id == eid,
        models.FDCSensor.name == _s(r["name"])).first()
    data = dict(unit=_s(r.get("unit")), warn_low=_f(r.get("warn_low")),
                warn_high=_f(r.get("warn_high")), alarm_low=_f(r.get("alarm_low")),
                alarm_high=_f(r.get("alarm_high")))
    if s:
        for k, v in data.items():
            setattr(s, k, v)
        return "updated"
    db.add(models.FDCSensor(equipment_id=eid, name=_s(r["name"]), **data))
    return "created"


spec("fdc_sensors", "FDC 센서", "설비×센서명 기준 생성/갱신 — 임계값 일괄 관리",
     [("asset_code", True, "설비 자산코드", "STK-2000-001"),
      ("name", True, "센서명", "권상모터 진동"),
      ("unit", False, "단위", "mm/s"),
      ("warn_low", False, "워닝 하한", ""),
      ("warn_high", False, "워닝 상한", "4.5"),
      ("alarm_low", False, "알람 하한", ""),
      ("alarm_high", False, "알람 상한", "7.1")],
     _imp_sensor,
     lambda db: [[s.equipment.asset_code, s.name, s.unit, s.warn_low, s.warn_high,
                  s.alarm_low, s.alarm_high]
                 for s in db.query(models.FDCSensor).all()])


# 8) L&L (추가 전용 — 자동 전파)
def _imp_lesson(db, ref, r):
    from .lessons import create_lesson_with_deployments
    create_lesson_with_deployments(
        db, title=_s(r["title"]), category=_s(r.get("category")).upper() or "DOWNTIME",
        model_id=ref.model(r.get("model_code"), required=False),
        problem=_s(r["problem"]), root_cause=_s(r.get("root_cause")),
        countermeasure=_s(r.get("countermeasure")),
        origin_site_id=ref.site(r["origin_site_code"]),
        created_by=_s(r.get("created_by")))
    return "created"


spec("lessons", "Lesson & Learn", "추가 전용 — 등록 시 전 사이트 자동 전파",
     [("title", True, "제목", "OO 베어링 조기 마모 — 주입주기 단축"),
      ("category", False, "SAFETY/QUALITY/DOWNTIME/COST", "DOWNTIME"),
      ("model_code", False, "관련 모델 코드 (공통이면 비움)", "STK-2000"),
      ("problem", True, "문제", "..."),
      ("root_cause", False, "근본원인", "..."),
      ("countermeasure", False, "대책", "..."),
      ("origin_site_code", True, "발생 사이트 코드", "KR1"),
      ("created_by", False, "작성자", "최정비")],
     _imp_lesson,
     lambda db: [[l.title, l.category,
                  l.model_id and db.get(models.EquipmentModel, l.model_id).code,
                  l.problem, l.root_cause, l.countermeasure, l.origin_site.code, l.created_by]
                 for l in db.query(models.Lesson).all()],
     insert_only=True)


# 9) 프로젝트
def _imp_project(db, ref, r):
    p = db.query(models.Project).filter(models.Project.code == _s(r["code"])).first()
    data = dict(name=_s(r["name"]), site_id=ref.site(r.get("site_code"), required=False),
                status=_s(r.get("status")).upper() or "PLANNING",
                budget=_f(r.get("budget")) or 0, owner=_s(r.get("owner")),
                start_date=_d(r.get("start_date")), end_date=_d(r.get("end_date")),
                description=_s(r.get("description")))
    if p:
        for k, v in data.items():
            setattr(p, k, v)
        return "updated"
    db.add(models.Project(code=_s(r["code"]), **data))
    return "created"


spec("projects", "프로젝트", "프로젝트 코드 기준 생성/갱신",
     [("code", True, "프로젝트 코드 (고유)", "P-2026-003"),
      ("name", True, "프로젝트명", "VN1 컨베이어 증설"),
      ("site_code", False, "사이트 코드", "VN1"),
      ("status", False, "PLANNING/ONGOING/DONE/HOLD", "PLANNING"),
      ("budget", False, "예산(원)", "550000000"),
      ("owner", False, "담당 PM", "김투자"),
      ("start_date", False, "착수일 YYYY-MM-DD", "2026-08-01"),
      ("end_date", False, "완료 목표일", "2027-02-28"),
      ("description", False, "설명", "")],
     _imp_project,
     lambda db: [[p.code, p.name, p.site.code if p.site else "", p.status, p.budget,
                  p.owner, p.start_date, p.end_date, p.description]
                 for p in db.query(models.Project).all()])


# ───────────────────────── 엔드포인트 ─────────────────────────

@router.get("/entities")
def entities():
    return [{"key": k, "label": v["label"], "desc": v["desc"],
             "insert_only": v["insert_only"],
             "columns": [{"name": c[0], "required": c[1], "desc": c[2]} for c in v["columns"]]}
            for k, v in SPECS.items()]


def _wb_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"})


def _styled_header(ws, columns):
    from openpyxl.styles import Font, PatternFill
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col[0] + ("*" if col[1] else ""))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1C6FD4" if col[1] else "8492A6")
        ws.column_dimensions[cell.column_letter].width = max(14, len(col[0]) + 4)


@router.get("/template/{entity}")
def template(entity: str):
    s = SPECS.get(entity)
    if not s:
        raise HTTPException(404, f"지원 항목: {list(SPECS)}")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"
    _styled_header(ws, s["columns"])
    for ci, col in enumerate(s["columns"], start=1):
        ws.cell(row=2, column=ci, value=col[3])  # 예시 행
    guide = wb.create_sheet("작성안내")
    guide.append(["컬럼", "필수", "설명", "예시"])
    for col in s["columns"]:
        guide.append([col[0], "필수" if col[1] else "", col[2], col[3]])
    guide.append([])
    guide.append([f"※ '데이터' 시트의 2행(예시)을 지우고 실제 데이터를 입력 후 업로드하세요."])
    guide.append([f"※ {s['desc']}"])
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["C"].width = 50
    return _wb_response(wb, f"양식_{s['label']}.xlsx")


@router.get("/export/{entity}")
def export(entity: str, db: Session = Depends(get_db)):
    s = SPECS.get(entity)
    if not s:
        raise HTTPException(404, f"지원 항목: {list(SPECS)}")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"
    _styled_header(ws, s["columns"])
    for row in s["exporter"](db):
        ws.append(["" if v is None else (v.isoformat() if isinstance(v, (date, datetime)) else v)
                   for v in row])
    return _wb_response(wb, f"{s['label']}_{date.today().isoformat()}.xlsx")


@router.post("/import/{entity}")
async def import_excel(entity: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    s = SPECS.get(entity)
    if not s:
        raise HTTPException(404, f"지원 항목: {list(SPECS)}")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "xlsx 파일을 열 수 없습니다 — 양식 파일을 사용하세요")
    ws = wb["데이터"] if "데이터" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "빈 시트입니다")
    headers = [str(h or "").strip().rstrip("*") for h in rows[0]]
    col_names = [c[0] for c in s["columns"]]
    missing = [c[0] for c in s["columns"] if c[1] and c[0] not in headers]
    if missing:
        raise HTTPException(400, f"필수 컬럼 누락: {missing} — 양식 파일을 사용하세요")

    ref = Ref(db)
    created = updated = 0
    errors = []
    example_values = {c[0]: c[3] for c in s["columns"]}
    for rno, raw in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {h: raw[i] if i < len(raw) else None for i, h in enumerate(headers) if h}
        # 예시 행 그대로면 건너뜀
        if all(_s(row.get(k)) == _s(v) for k, v in example_values.items() if v):
            continue
        try:
            for cname, required, *_ in s["columns"]:
                if required and not _s(row.get(cname)):
                    raise ValueError(f"필수값 누락: {cname}")
            result = s["importer"](db, ref, row)
            if result == "created":
                created += 1
            else:
                updated += 1
        except (ValueError, KeyError) as e:
            errors.append({"row": rno, "error": str(e)})
        except Exception as e:
            errors.append({"row": rno, "error": f"형식 오류: {e}"})
    if errors and created + updated == 0:
        db.rollback()
    else:
        db.commit()
    return {"entity": entity, "created": created, "updated": updated,
            "errors": errors[:30], "error_count": len(errors)}
