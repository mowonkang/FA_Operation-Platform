"""견적서 자동 데이터 분석.

CSV/XLSX 업로드 → 컬럼 자동 인식 → 항목 자동 분류 → 원가구조/이상단가/계산오류/중복 분석.
"""
import csv
import io
import math
import re
import statistics

CATEGORY_LABEL = {
    "MECH": "기구", "DRIVE": "구동", "ELEC": "전장", "CONTROL": "제어",
    "SW": "SW", "INSTALL": "설치/공사", "ETC": "기타",
}

# 항목명 키워드 → 카테고리 자동 분류 (순서 = 우선순위)
CATEGORY_KEYWORDS: list[tuple[str, list[str]]] = [
    ("SW", ["소프트웨어", "software", "라이선스", "license", "프로그램", "mes", "wcs", "wms"]),
    ("CONTROL", ["plc", "서보", "servo", "인버터", "inverter", "드라이브", "drive", "hmi",
                 "컨트롤", "controller", "제어", "모션", "엔코더", "encoder", "io카드", "i/o"]),
    ("DRIVE", ["모터", "motor", "감속기", "reducer", "기어", "gear", "베어링", "bearing",
               "휠", "wheel", "체인", "chain", "벨트", "belt", "로프", "rope", "풀리",
               "스프로켓", "sprocket", "커플링", "coupling", "lm가이드", "볼스크류"]),
    ("ELEC", ["센서", "sensor", "케이블", "cable", "판넬", "panel", "배선", "전장",
              "스위치", "switch", "릴레이", "relay", "전원", "smps", "차단기", "단자",
              "라이트커튼", "스캐너", "scanner", "포토", "근접"]),
    ("INSTALL", ["설치", "시운전", "운송", "포장", "공사", "인건", "출장", "교육",
                 "install", "setup", "commissioning", "기술료", "감리"]),
    ("MECH", ["프레임", "frame", "가공", "브라켓", "bracket", "레일", "rail", "판금",
              "용접", "구조", "마스트", "캐리지", "포크", "샤프트", "shaft", "플레이트", "커버"]),
]

HEADER_HINTS = {
    "name": ["품명", "항목", "품목", "내역", "name", "item", "description", "자재명"],
    "spec": ["규격", "사양", "spec", "model", "형식"],
    "qty": ["수량", "qty", "q'ty", "quantity", "수 량"],
    "unit_price": ["단가", "unit price", "unitprice", "단 가"],
    "amount": ["금액", "amount", "합계", "총액", "price", "금 액"],
    "remark": ["비고", "remark", "note"],
}


def classify(name: str) -> str:
    low = (name or "").lower()
    for cat, kws in CATEGORY_KEYWORDS:
        if any(k in low for k in kws):
            return cat
    return "ETC"


def _to_num(v) -> float:
    if v is None:
        return 0.0
    s = re.sub(r"[,\s원₩$]", "", str(v))
    if not s or s in ("-",):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _detect_columns(header: list[str]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for i, cell in enumerate(header):
        low = str(cell or "").strip().lower()
        for key, hints in HEADER_HINTS.items():
            if key not in cols and any(h in low for h in hints):
                cols[key] = i
    return cols


def parse_file(data: bytes, filename: str) -> list[dict]:
    """CSV 또는 XLSX 를 항목 리스트로 파싱. 헤더 행을 자동 탐색한다."""
    rows: list[list]
    if filename.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    else:
        text = None
        for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
            try:
                text = data.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("파일 인코딩을 인식할 수 없습니다 (UTF-8 또는 CP949 사용)")
        rows = list(csv.reader(io.StringIO(text)))

    header_idx, cols = -1, {}
    for i, row in enumerate(rows[:20]):
        c = _detect_columns([str(x or "") for x in row])
        if "name" in c and ("amount" in c or ("qty" in c and "unit_price" in c)):
            header_idx, cols = i, c
            break
    if header_idx < 0:
        raise ValueError("헤더 행을 찾지 못했습니다 — 품명/수량/단가/금액 컬럼이 필요합니다")

    items = []
    for line_no, row in enumerate(rows[header_idx + 1:], start=1):
        get = lambda k: row[cols[k]] if k in cols and cols[k] < len(row) else None
        name = str(get("name") or "").strip()
        if not name or name.lower() in ("합계", "총계", "total", "소계", "subtotal", "vat", "부가세"):
            continue
        qty = _to_num(get("qty")) or 1
        unit_price = _to_num(get("unit_price"))
        amount = _to_num(get("amount")) or qty * unit_price
        if amount == 0 and unit_price == 0:
            continue
        if unit_price == 0 and qty:
            unit_price = amount / qty
        items.append({
            "line_no": line_no, "name": name, "spec": str(get("spec") or "").strip(),
            "category": classify(name), "qty": qty, "unit_price": unit_price,
            "amount": amount, "remark": str(get("remark") or "").strip(),
        })
    if not items:
        raise ValueError("유효한 견적 항목이 없습니다")
    return items


def _norm_name(name: str) -> str:
    return re.sub(r"[\s\-_/()]", "", name.lower())


def analyze(items: list[dict]) -> dict:
    """단일 견적 분석: 원가구조, 계산오류, 이상단가, 중복, 고액 항목."""
    total = sum(i["amount"] for i in items)

    by_cat: dict[str, dict] = {}
    for i in items:
        agg = by_cat.setdefault(i["category"], {"amount": 0.0, "count": 0})
        agg["amount"] += i["amount"]
        agg["count"] += 1
    structure = [
        {"category": c, "label": CATEGORY_LABEL.get(c, c), "amount": v["amount"],
         "count": v["count"], "pct": round(v["amount"] / total * 100, 1) if total else 0}
        for c, v in sorted(by_cat.items(), key=lambda x: -x[1]["amount"])
    ]

    calc_errors = []
    for i in items:
        expected = i["qty"] * i["unit_price"]
        if i["qty"] and i["unit_price"] and abs(expected - i["amount"]) > max(1.0, expected * 0.001):
            calc_errors.append({"name": i["name"], "qty": i["qty"], "unit_price": i["unit_price"],
                                "amount": i["amount"], "expected": round(expected, 2),
                                "diff": round(i["amount"] - expected, 2)})

    # 카테고리 내 단가 z-score 이상치 (3개 이상일 때)
    outliers = []
    for cat, _ in by_cat.items():
        prices = [i["unit_price"] for i in items if i["category"] == cat and i["unit_price"] > 0]
        if len(prices) < 3:
            continue
        mean = statistics.mean(prices)
        std = statistics.pstdev(prices)
        if std == 0:
            continue
        for i in items:
            if i["category"] == cat and i["unit_price"] > 0:
                z = (i["unit_price"] - mean) / std
                if z > 2.0:
                    outliers.append({"name": i["name"], "category": cat,
                                     "unit_price": i["unit_price"],
                                     "category_mean": round(mean), "z_score": round(z, 1)})

    names: dict[str, list] = {}
    for i in items:
        names.setdefault(_norm_name(i["name"]), []).append(i)
    duplicates = [
        {"name": v[0]["name"], "count": len(v), "total_amount": sum(x["amount"] for x in v)}
        for v in names.values() if len(v) >= 2
    ]

    top_items = sorted(items, key=lambda x: -x["amount"])[:8]
    top = [{"name": i["name"], "category": i["category"], "amount": i["amount"],
            "pct": round(i["amount"] / total * 100, 1) if total else 0} for i in top_items]

    # 누적 80% 도달 항목 수 (파레토)
    cum, pareto_n = 0.0, 0
    for i in sorted(items, key=lambda x: -x["amount"]):
        cum += i["amount"]
        pareto_n += 1
        if total and cum >= total * 0.8:
            break

    findings = []
    if calc_errors:
        findings.append(f"⚠ 계산오류 {len(calc_errors)}건 — 수량×단가 ≠ 금액 (Nego 전 정정 요청)")
    if duplicates:
        findings.append(f"⚠ 중복 의심 항목 {len(duplicates)}건 — 통합/삭제(Eliminate) 검토")
    if outliers:
        findings.append(f"⚠ 카테고리 평균 대비 고단가 항목 {len(outliers)}건 — 근거 요청 또는 Reduce 검토")
    inst = next((s for s in structure if s["category"] == "INSTALL"), None)
    if inst and inst["pct"] > 20:
        findings.append(f"설치/공사비 비중 {inst['pct']}% — 통상 10~15% 대비 높음, 내역 세분화 요청")
    sw = next((s for s in structure if s["category"] == "SW"), None)
    if sw and sw["pct"] > 15:
        findings.append(f"SW/라이선스 비중 {sw['pct']}% — 범위·기존 라이선스 재사용(ERRC: Eliminate) 검토")
    findings.append(f"파레토: 상위 {pareto_n}개 항목이 전체 금액의 80% — Nego 집중 대상")

    return {
        "total_amount": total,
        "item_count": len(items),
        "structure": structure,
        "calc_errors": calc_errors,
        "price_outliers": outliers,
        "duplicates": duplicates,
        "top_items": top,
        "pareto_top_n": pareto_n,
        "findings": findings,
    }


def compare(quotes: list[dict]) -> dict:
    """동일 프로젝트 복수 견적 비교: 총액, 카테고리 매트릭스, 공통 항목 단가차."""
    if len(quotes) < 2:
        return {"error": "비교에는 2개 이상의 견적이 필요합니다"}
    vendors = [q["vendor"] for q in quotes]
    totals = [{"vendor": q["vendor"], "total": q["total"], "item_count": len(q["items"])}
              for q in quotes]
    base = min(q["total"] for q in quotes)
    for t in totals:
        t["vs_lowest_pct"] = round((t["total"] - base) / base * 100, 1) if base else 0

    cats = sorted({i["category"] for q in quotes for i in q["items"]})
    matrix = []
    for c in cats:
        row = {"category": c, "label": CATEGORY_LABEL.get(c, c)}
        for q in quotes:
            row[q["vendor"]] = sum(i["amount"] for i in q["items"] if i["category"] == c)
        matrix.append(row)

    # 공통 항목 단가 비교 (정규화 명칭 매칭)
    maps = [{_norm_name(i["name"]): i for i in q["items"]} for q in quotes]
    common_keys = set(maps[0])
    for m in maps[1:]:
        common_keys &= set(m)
    common = []
    for k in common_keys:
        prices = [m[k]["unit_price"] for m in maps]
        lo, hi = min(prices), max(prices)
        common.append({
            "name": maps[0][k]["name"],
            "prices": {vendors[j]: prices[j] for j in range(len(vendors))},
            "spread_pct": round((hi - lo) / lo * 100, 1) if lo else 0,
            "best_vendor": vendors[prices.index(lo)],
        })
    common.sort(key=lambda x: -x["spread_pct"])

    nego_potential = sum(
        (max(c["prices"].values()) - min(c["prices"].values())) for c in common
    )
    return {
        "vendors": vendors,
        "totals": totals,
        "category_matrix": matrix,
        "common_items": common[:30],
        "common_item_count": len(common),
        "nego_potential_amount": round(nego_potential),
        "guide": "공통 항목 단가차(spread)가 큰 순 — 최저가 업체 단가를 협상 기준선으로 활용",
    }
